#!/usr/bin/env python3
"""Parse official NCES/IPEDS release metadata and table documentation."""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin

from openpyxl import load_workbook

NCES_IPEDS_ACCESS_PAGE = "https://nces.ed.gov/ipeds/use-the-data/download-access-database"
DATA_GENERATOR_URL = "https://nces.ed.gov/ipeds/data-generator?year={year}&tableName={table_name}&HasRV=0&type=csv"
MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


@dataclass(frozen=True)
class ReleaseLink:
    collection_year: str
    data_year: int
    release_type: str
    release_date: str | None
    access_url: str | None
    metadata_url: str


@dataclass(frozen=True)
class IpedsTable:
    table_name: str
    survey_component: str | None
    year_coverage: str | None
    table_number: int | None
    table_title: str | None
    description: str | None
    table_release: str | None
    table_release_date: str | None


@dataclass(frozen=True)
class IpedsColumn:
    table_name: str
    var_name: str
    survey_component: str | None
    table_number: int | None
    table_title: str | None
    var_number: int | None
    var_order: int | None
    imputation_var: str | None
    var_title: str | None
    data_type: str | None
    field_width: int | None
    format: str | None
    multi_record: bool | None
    has_rv: str | None
    file_number: int | None
    section_number: int | None
    long_description: str | None
    var_source: str | None
    file_title: str | None
    section_title: str | None


@dataclass(frozen=True)
class IpedsValueLabel:
    table_name: str
    var_name: str
    code_value: str
    value_label: str | None
    frequency: int | None
    percent: float | None
    value_order: int | None
    var_title: str | None


@dataclass(frozen=True)
class TablesDoc:
    tables: list[IpedsTable]
    columns: list[IpedsColumn]
    value_labels: list[IpedsValueLabel]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def release_type_from_text(value: str) -> str:
    normalized = value.lower()
    if "preliminary" in normalized:
        return "preliminary"
    if "provisional" in normalized:
        return "provisional"
    if "final" in normalized:
        return "final"
    return "provisional"


def normalize_release_date_text(value: str | None) -> tuple[str | None, str | None]:
    """Convert NCES release-date text into an ISO date and precision.

    The IPEDS Access Database page currently publishes month-level dates like
    "March 2026". Other NCES pages use day-level dates, so the normalizer
    supports both while preserving precision for provenance notes.
    """
    if value is None:
        return None, None
    text = value.strip()
    if not text:
        return None, None

    day_match = re.fullmatch(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),\s*(\d{4})",
        text,
        flags=re.I,
    )
    if day_match:
        month = MONTHS[day_match.group(1).lower()]
        day = int(day_match.group(2))
        year = int(day_match.group(3))
        return f"{year:04d}-{month:02d}-{day:02d}", "day"

    month_match = re.fullmatch(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
        text,
        flags=re.I,
    )
    if month_match:
        month = MONTHS[month_match.group(1).lower()]
        year = int(month_match.group(2))
        return f"{year:04d}-{month:02d}-01", "month"

    iso_match = re.fullmatch(r"\d{4}-\d{2}-\d{2}", text)
    if iso_match:
        return text, "day"

    return None, None


def parse_access_page(html: str, base_url: str = NCES_IPEDS_ACCESS_PAGE) -> list[ReleaseLink]:
    """Extract official release links from the NCES Access database page.

    The page exposes paired Access ZIP and Excel metadata links. We intentionally
    avoid scraping data values here; this only discovers official source URLs.
    """
    hrefs = re.findall(r'href=["\']([^"\']+)["\']', html, flags=re.I)
    metadata_by_collection: dict[str, str] = {}
    access_by_collection: dict[str, str] = {}
    for href in hrefs:
        absolute = urljoin(base_url, href)
        lower = absolute.lower()
        if "tablesdoc.xlsx" in lower:
            match = re.search(r"ipeds[_-]?(\d{4})(\d{2})?tablesdoc\.xlsx", lower, flags=re.I)
            if match:
                start = int(match.group(1))
                end = int(match.group(2) or ((start + 1) % 100))
                metadata_by_collection[f"{start}-{end:02d}"] = absolute
        elif "ipeds_" in lower and lower.endswith(".zip"):
            match = re.search(r"ipeds_(\d{4})-(\d{2})", lower, flags=re.I)
            if match:
                access_by_collection[f"{int(match.group(1))}-{int(match.group(2)):02d}"] = absolute

    text = re.sub(r"\s+", " ", html)
    releases: list[ReleaseLink] = []
    for collection_year, metadata_url in sorted(metadata_by_collection.items(), reverse=True):
        start_year = int(collection_year[:4])
        access_url = access_by_collection.get(collection_year)
        row_match = re.search(
            rf"{re.escape(collection_year)}\s+Access.*?"
            rf"{re.escape(collection_year)}\s+Excel.*?"
            rf"(Preliminary|Provisional|Final)\s+"
            rf"((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{{4}})",
            text,
            flags=re.I,
        )
        window = row_match.group(0) if row_match else text[max(0, text.find(collection_year) - 400) : text.find(collection_year) + 800]
        date_match = re.search(
            r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}",
            window,
        )
        access_lower = (access_url or "").lower()
        release_type_text = access_url if any(token in access_lower for token in ("final", "provisional", "preliminary")) else window
        releases.append(
            ReleaseLink(
                collection_year=collection_year,
                data_year=start_year,
                release_type=release_type_from_text(release_type_text),
                release_date=date_match.group(0) if date_match else None,
                access_url=access_url,
                metadata_url=metadata_url,
            )
        )
    return releases


def parse_tablesdoc(path: Path) -> TablesDoc:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        tables = [_table_from_row(row) for row in _sheet_dicts(workbook, "tables")]
        columns = [_column_from_row(row) for row in _sheet_dicts(workbook, "vartable")]
        labels = [_label_from_row(row) for row in _sheet_dicts(workbook, "valuesets")]
    finally:
        workbook.close()
    return TablesDoc(
        tables=[table for table in tables if table.table_name],
        columns=[column for column in columns if column.table_name and column.var_name],
        value_labels=[label for label in labels if label.table_name and label.var_name and label.code_value != ""],
    )


def _sheet_dicts(workbook: Any, prefix: str) -> Iterable[dict[str, Any]]:
    sheet_name = next((name for name in workbook.sheetnames if name.lower().startswith(prefix.lower())), None)
    if sheet_name is None:
        raise ValueError(f"Workbook is missing a {prefix} sheet")
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    for row in rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        yield {headers[i].lower(): row[i] if i < len(row) else None for i in range(len(headers))}


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _bool(value: Any) -> bool | None:
    text = _text(value)
    if text is None:
        return None
    if text.lower() in {"1", "true", "yes", "y"}:
        return True
    if text.lower() in {"0", "false", "no", "n"}:
        return False
    return None


def _table_from_row(row: dict[str, Any]) -> IpedsTable:
    return IpedsTable(
        table_name=_text(row.get("tablename")) or "",
        survey_component=_text(row.get("survey")),
        year_coverage=_text(row.get("yearcoverage")),
        table_number=_int(row.get("tablenumber")),
        table_title=_text(row.get("tabletitle")),
        description=_text(row.get("description")),
        table_release=_text(row.get("release")),
        table_release_date=_text(row.get("release_date")),
    )


def _column_from_row(row: dict[str, Any]) -> IpedsColumn:
    return IpedsColumn(
        table_name=_text(row.get("tablename")) or "",
        var_name=(_text(row.get("varname")) or "").upper(),
        survey_component=_text(row.get("survey")),
        table_number=_int(row.get("tablenumber")),
        table_title=_text(row.get("tabletitle")),
        var_number=_int(row.get("varnumber")),
        var_order=_int(row.get("varorder")),
        imputation_var=(_text(row.get("imputationvar")) or "").upper() or None,
        var_title=_text(row.get("vartitle")),
        data_type=_text(row.get("datatype")),
        field_width=_int(row.get("fieldwidth")),
        format=_text(row.get("format")),
        multi_record=_bool(row.get("multirecord")),
        has_rv=_text(row.get("hasrv")),
        file_number=_int(row.get("filenumber")),
        section_number=_int(row.get("sectionnumber")),
        long_description=_text(row.get("longdescription")),
        var_source=_text(row.get("varsource")),
        file_title=_text(row.get("filetitle")),
        section_title=_text(row.get("sectiontitle")),
    )


def _label_from_row(row: dict[str, Any]) -> IpedsValueLabel:
    return IpedsValueLabel(
        table_name=_text(row.get("tablename")) or "",
        var_name=(_text(row.get("varname")) or "").upper(),
        code_value=_text(row.get("codevalue")) or "",
        value_label=_text(row.get("valuelabel")),
        frequency=_int(row.get("frequency")),
        percent=_float(row.get("percent")),
        value_order=_int(row.get("valueorder")),
        var_title=_text(row.get("vartitle")),
    )
