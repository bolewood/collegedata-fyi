#!/usr/bin/env python3
"""Build web/public/recipes/acceptance-vs-yield-starter.xlsx.

Reads the checked-in College Pricing Power dataset
(web/src/lib/pricing-power-recipe-data.ts) so the workbook matches the
interactive page. There was no prior generator: the original XLSX was a
hand-filled 18-school CDS C1 template.

Rebuild after regenerating the TypeScript dataset:

    python3 tools/ipeds/build_pricing_power_recipe.py
    python3 tools/ipeds/build_pricing_power_starter_xlsx.py
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover - install hint
    raise SystemExit(
        "error: openpyxl is required. Install with "
        "`python3 -m pip install --user 'openpyxl>=3.1'` "
        "or `python3 -m pip install -r tools/ipeds/requirements.txt`"
    ) from exc

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DATA = REPO_ROOT / "web/src/lib/pricing-power-recipe-data.ts"
DEFAULT_OUTPUT = REPO_ROOT / "web/public/recipes/acceptance-vs-yield-starter.xlsx"

PANEL_A_HEADERS = (
    "school",
    "school_id",
    "ipeds_id",
    "applied",
    "admitted",
    "enrolled",
    "acceptance",
    "yield",
    "cds_acceptance",
    "cds_yield",
    "cds_year",
)
PANEL_B_HEADERS = PANEL_A_HEADERS + (
    "burden",
    "median_debt",
    "monthly_payment",
    "earnings_10yr",
    "avg_net_price",
    "instruction_fte",
    "instruction_net_price_ratio",
)

RATE_HEADERS = frozenset(
    {
        "acceptance",
        "yield",
        "cds_acceptance",
        "cds_yield",
        "burden",
        "instruction_net_price_ratio",
    }
)
COUNT_HEADERS = frozenset({"applied", "admitted", "enrolled"})
MONEY_HEADERS = frozenset(
    {"median_debt", "earnings_10yr", "avg_net_price", "instruction_fte"}
)
CENTS_HEADERS = frozenset({"monthly_payment"})

HEADER_FONT = Font(bold=True, name="Calibri", size=11)
HEADER_FILL = PatternFill("solid", fgColor="E8E4D9")
TITLE_FONT = Font(bold=True, name="Calibri", size=14)
SECTION_FONT = Font(bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def load_pricing_power_dataset(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Parse PRICING_POWER_META and PRICING_POWER_SCHOOLS from the generated TS."""
    text = path.read_text(encoding="utf-8")
    decoder = json.JSONDecoder()
    meta_marker = "export const PRICING_POWER_META ="
    schools_marker = "export const PRICING_POWER_SCHOOLS"
    try:
        meta_at = text.index(meta_marker)
        meta, _ = decoder.raw_decode(text, text.index("{", meta_at))
        schools_at = text.index(schools_marker)
        # Skip the `PricingPowerSchool[]` type annotation; the payload follows `= [`.
        schools, _ = decoder.raw_decode(text, text.index("= [", schools_at) + 2)
    except (ValueError, json.JSONDecodeError) as exc:
        raise ValueError(f"could not parse pricing-power dataset from {path}") from exc
    if not isinstance(meta, dict) or not isinstance(schools, list):
        raise ValueError(f"unexpected dataset shape in {path}")
    if any(not isinstance(row, dict) for row in schools):
        raise ValueError(f"pricing-power schools in {path} are not objects")
    return meta, schools


def is_panel_b(row: dict[str, Any]) -> bool:
    return all(
        row.get(key) is not None
        for key in (
            "burden",
            "medianDebt",
            "monthlyPayment",
            "earnings10yr",
            "avgNetPrice",
            "instructionFte",
            "instructionNetPriceRatio",
        )
    )


def _panel_a_values(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row.get("name"),
        row.get("schoolId"),
        str(row.get("ipedsId") or ""),
        row.get("applied"),
        row.get("admitted"),
        row.get("enrolled"),
        row.get("acceptanceRate"),
        row.get("yieldRate"),
        row.get("cdsAcceptanceRate"),
        row.get("cdsYieldRate"),
        row.get("cdsYear"),
    )


def _panel_b_values(row: dict[str, Any]) -> tuple[Any, ...]:
    return _panel_a_values(row) + (
        row.get("burden"),
        row.get("medianDebt"),
        row.get("monthlyPayment"),
        row.get("earnings10yr"),
        row.get("avgNetPrice"),
        row.get("instructionFte"),
        row.get("instructionNetPriceRatio"),
    )


def _style_header(ws: Worksheet, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _apply_number_formats(ws: Worksheet, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        if header in RATE_HEADERS:
            fmt: str = numbers.FORMAT_PERCENTAGE_00
        elif header in COUNT_HEADERS or header in MONEY_HEADERS:
            fmt = "#,##0"
        elif header in CENTS_HEADERS:
            fmt = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        else:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.number_format = fmt


def _autosize(ws: Worksheet, headers: tuple[str, ...], cap: int = 36) -> None:
    for col, header in enumerate(headers, start=1):
        width = len(header) + 2
        sample_rows = min(ws.max_row, 40)
        for row in range(2, sample_rows + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            width = max(width, min(len(str(value)) + 2, cap))
        if header == "school":
            width = 42
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_sheet(
    ws: Worksheet,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    _style_header(ws, headers)
    _apply_number_formats(ws, headers)
    _autosize(ws, headers)


def _pct(value: float, digits: int = 2) -> str:
    return f"{value * 100:.{digits}f}%"


def _readme_rows(meta: dict[str, Any], panel_a_n: int, panel_b_n: int) -> list[tuple[str, str]]:
    exclusions = meta.get("exclusions") or {}
    scorecard_years = ", ".join(meta.get("scorecardYears") or []) or "2022-23"
    return [
        ("College Pricing Power", ""),
        (
            "Starter workbook for https://www.collegedata.fyi/recipes/acceptance-vs-yield",
            "",
        ),
        (
            "Generated from web/src/lib/pricing-power-recipe-data.ts. "
            "Do not edit by hand — regenerate this file.",
            "",
        ),
        ("", ""),
        ("Sources", ""),
        (
            "IPEDS ADM2024 raw counts",
            "applicants_total, admissions_total, enrolled_total from "
            "school_facts_unified (source_table=ADM2024). Rates are computed "
            "from those counts. Never use admit_rate_total / yield_rate_total "
            "(DRVADM2024 integer-rounded derived rates).",
        ),
        (
            "College Scorecard",
            f"{scorecard_years} file via scorecard_summary: earnings_10yr_median, "
            "median_debt_monthly_payment, median_debt_completers, avg_net_price, "
            "instructional_expenditure_fte.",
        ),
        (
            "CDS C1 cross-check",
            "school_browser_rows 2024-25, sub_institutional is null, complete "
            "acceptance_rate and yield_rate. Tooltip only — not plotted. "
            f"{meta.get('cdsAttachedCount')} of {meta.get('panelACount')} Panel A "
            f"schools ({meta.get('cdsCrosscheckCount')} complete C1 rows in the build).",
        ),
        ("", ""),
        ("Cycle dating (three clocks)", ""),
        (
            "Fig. 1 admissions",
            str(meta.get("ipedsCycle") or "fall 2024 (ADM2024)"),
        ),
        (
            "Fig. 2 outcomes",
            f"College Scorecard {scorecard_years} net price, debt, payments, "
            "earnings, instructional spending. Earnings describe federally "
            "aided students from a much earlier entering cohort (~10 years).",
        ),
        (
            "Not on the charts",
            "Wall Street Journal operational reporting on fall 2025 recruiting "
            "and the fall 2026 shortfall. The join is by institution, not by "
            "one class of students moving through college.",
        ),
        ("", ""),
        ("Formulas", ""),
        ("acceptance", "admitted / applied"),
        ("yield", "enrolled / admitted"),
        (
            "burden",
            "monthly_payment × 12 / earnings_10yr. "
            "Worked example: $3,000 / $60,000 = 5%.",
        ),
        (
            "instruction / net-price",
            "instruction_fte / avg_net_price. Not the share of a college's "
            "budget spent on teaching. Denominator is Title IV average net "
            "price, not total spending.",
        ),
        ("", ""),
        ("Sample", ""),
        (
            "Panel A",
            f"{panel_a_n} schools. Median acceptance {_pct(meta['medianAcceptance'])}, "
            f"median yield {_pct(meta['medianYield'])}.",
        ),
        (
            "Panel B",
            f"{panel_b_n} schools (Panel A plus positive Scorecard debt, "
            f"earnings, net price, and instruction). Median yield "
            f"{_pct(meta['medianYieldB'])}, median burden {_pct(meta['medianBurden'])}.",
        ),
        (
            "Exclusions (builder tallies)",
            "out of scope "
            f"{exclusions.get('outOfScope', 0)}; missing/zero ADM counts "
            f"{exclusions.get('missingZeroCounts', 0)}; admitted > applied "
            f"{exclusions.get('admittedGtApplied', 0)}; enrolled > admitted "
            f"{exclusions.get('enrolledGtAdmitted', 0)}; missing or non-positive "
            f"Scorecard fields {exclusions.get('missingNonpositiveScorecard', 0)}.",
        ),
        ("", ""),
        ("Limitations (one line each)", ""),
        (
            "Yield is not pure demand",
            "Early Decision, geography, aid, athletics, mission, and a "
            "self-selected pool can all produce the same conversion rate.",
        ),
        (
            "Title IV net price only",
            "Scorecard average net price covers undergraduates who received "
            "Title IV federal aid. Full-pay students are not in the average.",
        ),
        (
            "Federal debt only",
            "Burden uses estimated federal-loan payments. Families also pay "
            "with savings, income, grants, parent PLUS, and private loans.",
        ),
        (
            "Three clocks",
            "Fall 2024 admits, 2022-23 Scorecard outcomes, and later news "
            "events are not one cohort. Do not read the sheets as a single "
            "class moving through college.",
        ),
        (
            "Debt bunches at loan limits",
            "Median completer debt clusters at common federal limits "
            "(170 Panel B schools report exactly $27,000). Burden differences "
            "in the middle of the sample are often earnings differences.",
        ),
        (
            "Branch campuses share Scorecard records",
            "Some related campuses inherit a parent Scorecard row, so debt, "
            "earnings, and burden can repeat. Repeats are kept as reported.",
        ),
        (
            "Instruction / net-price is not a budget share",
            "Do not treat the remainder as administration. The two inputs "
            "come from different systems and populations.",
        ),
        (
            "Correlation is not causation",
            "This workbook does not estimate price elasticity, markups, or "
            "the price at which a class would fail to fill.",
        ),
        ("", ""),
        ("Rebuild", ""),
        (
            "Dataset",
            "python3 tools/ipeds/build_pricing_power_recipe.py",
        ),
        (
            "This workbook",
            "python3 tools/ipeds/build_pricing_power_starter_xlsx.py",
        ),
        (
            "Anon key",
            "https://www.collegedata.fyi/api — school_facts_unified is "
            "anon-readable; PostgREST caps responses at 1,000 rows, so page "
            "with limit=1000&offset=0, then offset=1000, then offset=2000.",
        ),
    ]


def write_workbook(
    meta: dict[str, Any],
    schools: list[dict[str, Any]],
    dest: Path,
) -> dict[str, int]:
    panel_a_rows = [_panel_a_values(row) for row in schools]
    panel_b_rows = [_panel_b_values(row) for row in schools if is_panel_b(row)]
    if len(panel_a_rows) != meta.get("panelACount"):
        raise ValueError(
            f"Panel A row count {len(panel_a_rows)} != meta.panelACount "
            f"{meta.get('panelACount')}"
        )
    if len(panel_b_rows) != meta.get("panelBCount"):
        raise ValueError(
            f"Panel B row count {len(panel_b_rows)} != meta.panelBCount "
            f"{meta.get('panelBCount')}"
        )

    wb = Workbook()
    readme = wb.active
    assert readme is not None
    readme.title = "README"
    for label, value in _readme_rows(meta, len(panel_a_rows), len(panel_b_rows)):
        readme.append([label, value])
    for row in readme.iter_rows(min_row=1, max_row=readme.max_row, max_col=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
    readme["A1"].font = TITLE_FONT
    for cell in readme["A"]:
        if cell.value in {
            "Sources",
            "Cycle dating (three clocks)",
            "Formulas",
            "Sample",
            "Limitations (one line each)",
            "Rebuild",
        }:
            cell.font = SECTION_FONT
    readme.column_dimensions["A"].width = 36
    readme.column_dimensions["B"].width = 92
    readme.row_dimensions[1].height = 22
    for row_idx in range(2, readme.max_row + 1):
        readme.row_dimensions[row_idx].height = 36
    readme.freeze_panes = "A2"

    panel_a = wb.create_sheet("Panel A")
    panel_b = wb.create_sheet("Panel B")
    _write_sheet(panel_a, PANEL_A_HEADERS, panel_a_rows)
    _write_sheet(panel_b, PANEL_B_HEADERS, panel_b_rows)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return {
        "readmeRows": readme.max_row,
        "panelARows": panel_a.max_row,
        "panelBRows": panel_b.max_row,
        "panelAData": len(panel_a_rows),
        "panelBData": len(panel_b_rows),
    }


def verify_workbook(path: Path, meta: dict[str, Any]) -> dict[str, int]:
    wb = load_workbook(path, read_only=True, data_only=False)
    names = set(wb.sheetnames)
    expected = {"README", "Panel A", "Panel B"}
    if names != expected:
        raise ValueError(f"{path} sheets {sorted(names)} != {sorted(expected)}")
    panel_a = wb["Panel A"]
    panel_b = wb["Panel B"]
    a_rows = panel_a.max_row or 0
    b_rows = panel_b.max_row or 0
    a_header = [cell.value for cell in next(panel_a.iter_rows(min_row=1, max_row=1))]
    b_header = [cell.value for cell in next(panel_b.iter_rows(min_row=1, max_row=1))]
    wb.close()
    if a_header != list(PANEL_A_HEADERS):
        raise ValueError(f"Panel A headers {a_header}")
    if b_header != list(PANEL_B_HEADERS):
        raise ValueError(f"Panel B headers {b_header}")
    if a_rows != 1 + int(meta["panelACount"]):
        raise ValueError(f"Panel A row count {a_rows}")
    if b_rows != 1 + int(meta["panelBCount"]):
        raise ValueError(f"Panel B row count {b_rows}")
    return {"panelARows": a_rows, "panelBRows": b_rows}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data", type=Path, default=DEFAULT_DATA)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    if not args.data.is_file():
        print(f"error: dataset not found: {args.data}", file=sys.stderr)
        return 2
    try:
        meta, schools = load_pricing_power_dataset(args.data)
        counts = write_workbook(meta, schools, args.out)
        verified = verify_workbook(args.out, meta)
    except (OSError, ValueError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    print(
        f"wrote {args.out} (README {counts['readmeRows']} rows, "
        f"Panel A {verified['panelARows']} rows / {counts['panelAData']} schools, "
        f"Panel B {verified['panelBRows']} rows / {counts['panelBData']} schools)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
