from __future__ import annotations

import json
import sys
import tempfile
import unittest
import urllib.error
import zipfile
from contextlib import ExitStack
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from tools.ipeds.download_release import (
    download,
    export_access_tables,
    is_f2_finance_table,
    main as download_release_main,
    release_uses_access_source,
    select_release,
    source_mode,
)
from tools.ipeds.metadata import IpedsTable, ReleaseLink, TablesDoc, normalize_release_date_text, parse_access_page, parse_tablesdoc


class IpedsMetadataTests(unittest.TestCase):
    def test_download_streams_with_timeout_and_removes_oversized_partial_file(self) -> None:
        class FakeResponse:
            def __init__(self, chunks: list[bytes]) -> None:
                self.chunks = chunks

            def __enter__(self) -> "FakeResponse":
                return self

            def __exit__(self, *_args: object) -> None:
                return None

            def read(self, _size: int) -> bytes:
                return self.chunks.pop(0) if self.chunks else b""

        class FakeOpener:
            def __init__(self, chunks: list[bytes]) -> None:
                self.chunks = chunks
                self.timeout: int | None = None

            def open(self, _request: object, *, timeout: int) -> FakeResponse:
                self.timeout = timeout
                return FakeResponse(self.chunks)

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "source.zip"
            opener = FakeOpener([b"abc", b"def"])
            download(opener, "https://example.test/source.zip", path)  # type: ignore[arg-type]
            self.assertEqual(path.read_bytes(), b"abcdef")
            self.assertEqual(opener.timeout, 120)

            with patch("tools.ipeds.download_release.MAX_DOWNLOAD_BYTES", 5):
                with self.assertRaisesRegex(ValueError, "exceeded"):
                    download(
                        FakeOpener([b"abc", b"def"]),  # type: ignore[arg-type]
                        "https://example.test/source.zip",
                        path,
                    )
            self.assertFalse(path.exists())

    def test_parse_access_page_finds_latest_excel_and_zip(self) -> None:
        html = """
        Final release data include revisions to provisional release data.
        <a href="/ipeds/tablefiles/zipfiles/IPEDS_2024-25_Provisional.zip">2024-25 Access</a>
        <a href="/ipeds/tablefiles/tableDocs/IPEDS202425Tablesdoc.xlsx">2024-25 Excel</a>
        (IPEDS202425Tablesdoc.xlsx, 1261kb) Provisional March 2026
        <a href="/ipeds/tablefiles/zipfiles/IPEDS_2023-24.zip">2023-24 Access</a>
        <a href="/ipeds/tablefiles/tableDocs/IPEDS202324Tablesdoc.xlsx">2023-24 Excel</a>
        (IPEDS202324Tablesdoc.xlsx, 1348kb) Final March 2026
        """
        releases = parse_access_page(html)
        self.assertEqual(releases[0].collection_year, "2024-25")
        self.assertEqual(releases[0].data_year, 2024)
        self.assertEqual(releases[0].release_type, "provisional")
        self.assertEqual(releases[0].release_date, "March 2026")
        self.assertTrue(releases[0].metadata_url.endswith("IPEDS202425Tablesdoc.xlsx"))
        self.assertTrue(releases[0].access_url.endswith("IPEDS_2024-25_Provisional.zip"))
        self.assertEqual(releases[1].release_type, "final")

    def test_normalize_release_date_text_handles_month_and_day_precision(self) -> None:
        self.assertEqual(normalize_release_date_text("March 2026"), ("2026-03-01", "month"))
        self.assertEqual(normalize_release_date_text("January 6, 2026"), ("2026-01-06", "day"))
        self.assertEqual(normalize_release_date_text("2026-03-01"), ("2026-03-01", "day"))

    def test_download_release_selection_is_strict(self) -> None:
        releases = [
            ReleaseLink("2024-25", 2024, "provisional", "March 2026", "access.zip", "tables.xlsx"),
            ReleaseLink("2023-24", 2023, "final", "March 2026", "access-final.zip", "tables-final.xlsx"),
        ]

        self.assertEqual(select_release(releases, None).collection_year, "2024-25")
        self.assertEqual(select_release(releases, None, "final").collection_year, "2023-24")
        self.assertEqual(select_release(releases, "2023-24").data_year, 2023)
        self.assertEqual(select_release(releases, "2023-24", "final").release_type, "final")
        with self.assertRaisesRegex(SystemExit, "No IPEDS release matched 2022-23"):
            select_release(releases, "2022-23")
        with self.assertRaisesRegex(SystemExit, "No IPEDS release matched 2023-24 provisional"):
            select_release(releases, "2023-24", "provisional")

    def test_finance_table_detection_is_specific_to_f2(self) -> None:
        self.assertTrue(is_f2_finance_table("F2324_F2"))
        self.assertFalse(is_f2_finance_table("F2324_F1A"))
        self.assertFalse(is_f2_finance_table("SFA2324"))

    def test_final_release_routes_to_revised_access_source(self) -> None:
        final = ReleaseLink("2023-24", 2023, "final", "March 2026", "access.zip", "tables.xlsx")
        provisional = ReleaseLink("2024-25", 2024, "provisional", "March 2026", "access.zip", "tables.xlsx")

        self.assertTrue(release_uses_access_source(final))
        self.assertFalse(release_uses_access_source(provisional))
        self.assertEqual(source_mode([], ["F2223_F2"]), "access")
        self.assertEqual(source_mode(["HD2023"], ["F2223_F2"]), "mixed")
        self.assertEqual(source_mode([], []), "data_generator")

    def test_access_export_requires_an_official_bundle_url(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            exported = export_access_tables(
                opener=None,  # type: ignore[arg-type]
                access_url=None,
                out_dir=Path(tmp),
                table_names=["F2223_F2"],
            )

        self.assertEqual(exported, [])

    def test_access_export_handles_missing_tool_bad_bundle_and_table_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with patch("tools.ipeds.download_release.shutil.which", return_value=None):
                self.assertEqual(
                    export_access_tables(
                        opener=None,  # type: ignore[arg-type]
                        access_url="https://example.test/access.zip",
                        out_dir=out_dir,
                        table_names=["F2223_F2"],
                    ),
                    [],
                )

            def write_bundle(_opener: object, _url: str, path: Path) -> None:
                with zipfile.ZipFile(path, "w") as bundle:
                    bundle.writestr("readme.txt", "no database here")

            with (
                patch("tools.ipeds.download_release.shutil.which", return_value="/usr/bin/mdb-export"),
                patch("tools.ipeds.download_release.download", side_effect=write_bundle),
            ):
                self.assertEqual(
                    export_access_tables(
                        opener=None,  # type: ignore[arg-type]
                        access_url="https://example.test/access.zip",
                        out_dir=out_dir,
                        table_names=["F2223_F2"],
                    ),
                    [],
                )

            def write_access_bundle(_opener: object, _url: str, path: Path) -> None:
                with zipfile.ZipFile(path, "w") as bundle:
                    bundle.writestr("release.accdb", "placeholder")

            with (
                patch("tools.ipeds.download_release.shutil.which", return_value="/usr/bin/mdb-export"),
                patch("tools.ipeds.download_release.download", side_effect=write_access_bundle),
                patch("tools.ipeds.download_release.MAX_ACCESS_DB_BYTES", 5),
            ):
                with self.assertRaisesRegex(ValueError, "Access database member exceeds"):
                    export_access_tables(
                        opener=None,  # type: ignore[arg-type]
                        access_url="https://example.test/access.zip",
                        out_dir=out_dir,
                        table_names=["F2223_F2"],
                    )

            with (
                patch("tools.ipeds.download_release.shutil.which", return_value="/usr/bin/mdb-export"),
                patch("tools.ipeds.download_release.download", side_effect=write_access_bundle),
                patch("tools.ipeds.download_release.subprocess_run", side_effect=[0, 1]),
            ):
                (out_dir / "F2223_F2.zip").write_bytes(b"stale provisional")
                self.assertEqual(
                    export_access_tables(
                        opener=None,  # type: ignore[arg-type]
                        access_url="https://example.test/access.zip",
                        out_dir=out_dir,
                        table_names=["F2223_F2", "F2223_F1A"],
                    ),
                    ["F2223_F2"],
                )
                self.assertTrue((out_dir / "F2223_F2.csv").exists())
                self.assertFalse((out_dir / "F2223_F2.zip").exists())
                self.assertFalse((out_dir / "F2223_F1A.csv").exists())

    def test_download_main_routes_final_to_access_and_records_manifest(self) -> None:
        class FakeResponse:
            def read(self) -> bytes:
                return b"release page"

        class FakeOpener:
            def __init__(self) -> None:
                self.urls: list[str] = []

            def open(self, url: str) -> FakeResponse:
                self.urls.append(url)
                return FakeResponse()

        release = ReleaseLink(
            "2023-24",
            2023,
            "final",
            "March 2026",
            "https://example.test/access.zip",
            "https://example.test/tables.xlsx",
        )
        tablesdoc = TablesDoc(
            tables=[IpedsTable("F2223_F2", "Finance", "2022-23", 1, "Finance", None, None, None)],
            columns=[],
            value_labels=[],
        )
        opener = FakeOpener()
        with tempfile.TemporaryDirectory() as tmp:
            with ExitStack() as stack:
                stack.enter_context(patch.object(sys, "argv", [
                "download_release.py",
                "--collection-year", "2023-24",
                "--release-type", "final",
                "--out-dir", tmp,
                "--tables", "F2223_F2",
                ]))
                stack.enter_context(patch("tools.ipeds.download_release.build_opener", return_value=opener))
                stack.enter_context(patch("tools.ipeds.download_release.parse_access_page", return_value=[release]))
                stack.enter_context(patch("tools.ipeds.download_release.parse_tablesdoc", return_value=tablesdoc))
                download_mock = stack.enter_context(patch("tools.ipeds.download_release.download"))
                export_mock = stack.enter_context(
                    patch("tools.ipeds.download_release.export_access_tables", return_value=["F2223_F2"])
                )
                self.assertEqual(download_release_main(), 0)
                manifest = json.loads(
                    (Path(tmp) / "2023-24-final" / "release.json").read_text(encoding="utf-8")
                )

        self.assertEqual(download_mock.call_count, 1)
        export_mock.assert_called_once()
        self.assertEqual(manifest["source_mode"], "access")
        self.assertEqual(manifest["access_exported_tables"], ["F2223_F2"])
        self.assertEqual(opener.urls, ["https://nces.ed.gov/ipeds/use-the-data/download-access-database"])

    def test_download_main_uses_access_fallback_after_generator_404(self) -> None:
        class FakeResponse:
            def read(self) -> bytes:
                return b"release page"

        class FakeOpener:
            def open(self, _url: str) -> FakeResponse:
                return FakeResponse()

        release = ReleaseLink(
            "2024-25",
            2024,
            "provisional",
            "March 2026",
            "https://example.test/access.zip",
            "https://example.test/tables.xlsx",
        )
        tablesdoc = TablesDoc(
            tables=[IpedsTable("F2324_F2", "Finance", "2023-24", 1, "Finance", None, None, None)],
            columns=[],
            value_labels=[],
        )

        def generator_404(_opener: object, url: str, _path: Path) -> None:
            if "tableName=" in url:
                raise urllib.error.HTTPError(url, 404, "not found", None, None)

        with tempfile.TemporaryDirectory() as tmp:
            with ExitStack() as stack:
                stack.enter_context(patch.object(sys, "argv", [
                "download_release.py",
                "--collection-year", "2024-25",
                "--release-type", "provisional",
                "--out-dir", tmp,
                "--tables", "F2324_F2",
                "--access-fallback",
                ]))
                stack.enter_context(patch("tools.ipeds.download_release.build_opener", return_value=FakeOpener()))
                stack.enter_context(patch("tools.ipeds.download_release.parse_access_page", return_value=[release]))
                stack.enter_context(patch("tools.ipeds.download_release.parse_tablesdoc", return_value=tablesdoc))
                stack.enter_context(patch("tools.ipeds.download_release.download", side_effect=generator_404))
                export_mock = stack.enter_context(
                    patch("tools.ipeds.download_release.export_access_tables", return_value=["F2324_F2"])
                )
                self.assertEqual(download_release_main(), 0)
                manifest = json.loads(
                    (Path(tmp) / "2024-25-provisional" / "release.json").read_text(encoding="utf-8")
                )

        export_mock.assert_called_once()
        self.assertEqual(manifest["source_mode"], "access")
        self.assertEqual(manifest["unresolved_tables"], [])

    def test_download_main_rejects_incomplete_access_and_unresolved_generator(self) -> None:
        class FakeResponse:
            def read(self) -> bytes:
                return b"release page"

        class FakeOpener:
            def open(self, _url: str) -> FakeResponse:
                return FakeResponse()

        tablesdoc = TablesDoc(
            tables=[IpedsTable("F2223_F2", "Finance", "2022-23", 1, "Finance", None, None, None)],
            columns=[],
            value_labels=[],
        )
        final = ReleaseLink(
            "2023-24",
            2023,
            "final",
            "March 2026",
            "https://example.test/access.zip",
            "https://example.test/tables.xlsx",
        )
        with tempfile.TemporaryDirectory() as tmp, ExitStack() as stack:
            stack.enter_context(patch.object(sys, "argv", [
                "download_release.py",
                "--collection-year", "2023-24",
                "--release-type", "final",
                "--out-dir", tmp,
                "--tables", "F2223_F2",
            ]))
            stack.enter_context(patch("tools.ipeds.download_release.build_opener", return_value=FakeOpener()))
            stack.enter_context(patch("tools.ipeds.download_release.parse_access_page", return_value=[final]))
            stack.enter_context(patch("tools.ipeds.download_release.parse_tablesdoc", return_value=tablesdoc))
            stack.enter_context(patch("tools.ipeds.download_release.download"))
            stack.enter_context(patch("tools.ipeds.download_release.export_access_tables", return_value=[]))
            with self.assertRaisesRegex(SystemExit, "could not provide expected table"):
                download_release_main()
            self.assertFalse((Path(tmp) / "2023-24-final" / "release.json").exists())

        provisional = ReleaseLink(
            "2023-24",
            2023,
            "provisional",
            "March 2026",
            "https://example.test/access.zip",
            "https://example.test/tables.xlsx",
        )

        def generator_404(_opener: object, url: str, _path: Path) -> None:
            if "tableName=" in url:
                raise urllib.error.HTTPError(url, 404, "not found", None, None)

        with tempfile.TemporaryDirectory() as tmp, ExitStack() as stack:
            stack.enter_context(patch.object(sys, "argv", [
                "download_release.py",
                "--collection-year", "2023-24",
                "--release-type", "provisional",
                "--out-dir", tmp,
                "--tables", "F2223_F2",
            ]))
            stack.enter_context(patch("tools.ipeds.download_release.build_opener", return_value=FakeOpener()))
            stack.enter_context(patch("tools.ipeds.download_release.parse_access_page", return_value=[provisional]))
            stack.enter_context(patch("tools.ipeds.download_release.parse_tablesdoc", return_value=tablesdoc))
            stack.enter_context(patch("tools.ipeds.download_release.download", side_effect=generator_404))
            self.assertEqual(download_release_main(), 2)
            manifest = json.loads(
                (Path(tmp) / "2023-24-provisional" / "release.json").read_text(encoding="utf-8")
            )
            self.assertEqual(manifest["unresolved_tables"], ["F2223_F2"])

    def test_download_main_rejects_wrong_data_year_and_missing_finance_metadata(self) -> None:
        class FakeResponse:
            def read(self) -> bytes:
                return b"release page"

        class FakeOpener:
            def open(self, _url: str) -> FakeResponse:
                return FakeResponse()

        release = ReleaseLink(
            "2023-24",
            2023,
            "final",
            "March 2026",
            "https://example.test/access.zip",
            "https://example.test/tables.xlsx",
        )
        with ExitStack() as stack:
            stack.enter_context(patch.object(sys, "argv", [
                "download_release.py",
                "--collection-year", "2023-24",
                "--release-type", "final",
                "--data-year", "2022",
            ]))
            stack.enter_context(patch("tools.ipeds.download_release.build_opener", return_value=FakeOpener()))
            stack.enter_context(patch("tools.ipeds.download_release.parse_access_page", return_value=[release]))
            with self.assertRaisesRegex(SystemExit, "does not match collection year"):
                download_release_main()

        with tempfile.TemporaryDirectory() as tmp, ExitStack() as stack:
            release_dir = Path(tmp) / "2023-24-final"
            release_dir.mkdir()
            stale_manifest = release_dir / "release.json"
            stale_manifest.write_text('{"release_type":"provisional"}', encoding="utf-8")
            stack.enter_context(patch.object(sys, "argv", [
                "download_release.py",
                "--collection-year", "2023-24",
                "--release-type", "final",
                "--out-dir", tmp,
                "--tables", "F2223_F2",
            ]))
            stack.enter_context(patch("tools.ipeds.download_release.build_opener", return_value=FakeOpener()))
            stack.enter_context(patch("tools.ipeds.download_release.parse_access_page", return_value=[release]))
            stack.enter_context(
                patch(
                    "tools.ipeds.download_release.parse_tablesdoc",
                    return_value=TablesDoc(tables=[], columns=[], value_labels=[]),
                )
            )
            stack.enter_context(patch("tools.ipeds.download_release.download"))
            with self.assertRaisesRegex(SystemExit, "absent from the official Tablesdoc"):
                download_release_main()
            self.assertFalse(stale_manifest.exists())

    def test_parse_tablesdoc_reads_core_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tablesdoc.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "tables24"
            ws.append(["Survey", "YearCoverage", "TableName", "Tablenumber", "TableTitle", "Description", "Release", "Release_date"])
            ws.append(["HD", "2024", "HD2024", 1, "Directory", "Header", "Provisional", "March 2026"])
            ws = workbook.create_sheet("varTable24")
            ws.append(["Survey", "TableNumber", "TableName", "TableTitle", "VarNumber", "VarOrder", "VarName", "ImputationVar", "VarTitle", "DataType", "FieldWidth", "Format", "MultiRecord", "HasRV", "FileNumber", "SectionNumber", "LongDescription", "VarSource", "FileTitle", "SectionTitle"])
            ws.append(["HD", 1, "HD2024", "Directory", 1, 1, "INSTNM", "", "Institution name", "Alpha", 100, "", "No", "No", 1, 1, "Name", "Source", "File", "Section"])
            ws = workbook.create_sheet("valueSets24")
            ws.append(["TableName", "VarName", "Codevalue", "Frequency", "Percent", "ValueOrder", "ValueLabel", "VarTitle"])
            ws.append(["HD2024", "CONTROL", "1", 10, 50.0, 1, "Public", "Control"])
            workbook.save(path)

            parsed = parse_tablesdoc(path)
            self.assertEqual(parsed.tables[0].table_name, "HD2024")
            self.assertEqual(parsed.columns[0].var_name, "INSTNM")
            self.assertEqual(parsed.value_labels[0].value_label, "Public")


if __name__ == "__main__":
    unittest.main()
