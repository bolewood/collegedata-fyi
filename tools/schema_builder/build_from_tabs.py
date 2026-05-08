"""
Build a STRUCTURAL CDS schema from the per-section tabs of the commondataset.org
Excel template (CDS-A through CDS-J).

This is different from the Answer Sheet-based schema (see build_from_xlsx.py).
The Answer Sheet is a machine-readable canonical spec. Older templates
(2019-20 through 2023-24) only have the per-section tabs, which are visual
layouts for humans filling in data. The 2024-25 template is an intermediate
format: it has canonical question numbers, but each per-section tab is a
row-based metadata table rather than a visual form layout.

This script walks the per-section tabs and extracts the schema structure:
  - Section (A, B, C, ...)
  - Subsection markers (A0, A1, B1, B2, C1, ...) from column A, or from
    canonical question numbers in row-based sheets
  - Question row labels from column B
  - Column headers (Males | Females | Unknown, Full-Time | Part-Time, etc.)
  - Answer cell references for provenance

Output is a `cds_schema_YYYY_YY.structural.json` file. It does NOT contain
canonical question numbers (A.001, B.101, etc.) when the source sheet provides
them. Visual-layout historical years do not include canonical IDs; canonical
ID assignment for those years is a separate cross-reference step.

Structural schemas enable:
  - Per-year extraction that respects that year's specific field set
  - A cross-year diff tool (backlog P1) that flags additions/removals/renames
  - Cleaner authors targeting a specific year's template

Usage:
    python tools/schema_builder/build_from_tabs.py \\
        scratch/cds_templates_historical/CDS_2023-2024.xlsx \\
        schemas/cds_schema_2023_24.structural.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


# A subsection marker sits alone in column A. Matches the common CDS conventions:
#   A0, A1, B1, C21, H1, J1       (single-letter section + 1-3 digits)
#   I-1, I-2, H-1                 (letter-hyphen-digit, used in CDS-I + some others)
SUBSECTION_RE = re.compile(r"^([A-Z])-?(\d{1,3})\.?$")

# Lines that look like instruction bullets / prose, not field labels.
INSTRUCTION_PREFIXES = ("•", "*", "·", "Note:", "NOTE", "Provide", "Include",
                         "Complete", "Report", "Do not", "If your", "See ",
                         "For ", "Please", "http", "This ", "These ",
                         "In cases", "As used", "Nonresident - ",
                         "Eligible non-citizens",)

ROW_METADATA_REQUIRED_HEADERS = {
    "Question Number",
    "Question",
    "Answer",
    "Section",
    "Sub-Section",
}

ROW_METADATA_DIMENSION_HEADERS = {
    "Section": "section_title",
    "Sub-Section": "subsection_title",
    "Category": "category",
    "Student Group": "student_group",
    "Cohort": "cohort",
    "Residency": "residency",
    "Unit load": "unit_load",
    "Gender": "gender",
    "Value type": "value_type",
}


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    # Collapse internal newlines (cells with wrapped labels like
    # "Degree-Seeking\n First-Time\n First Year") into spaces.
    s = re.sub(r"\s+", " ", s)
    return s if s else None


def _looks_like_header_row(cells: list[str | None]) -> bool:
    """A row is a column-header row if col A + col B are empty (or col B is
    short-label-like) AND cols 2+ have at least 2 non-empty string cells, each
    of which is a short label (not a sentence)."""
    col_a, col_b = cells[0], cells[1]
    if col_a:
        return False
    # Col B can be empty (most headers) or a short row-group label like
    # "Undergraduate Students: Full-Time" that shares the header row.
    if col_b and (len(col_b) > 80 or col_b.endswith(".") or " " in col_b and len(col_b.split()) > 10):
        return False

    non_empty = [c for c in cells[2:] if c]
    if len(non_empty) < 2:
        return False

    # Each header cell should be a short label, not a formula or sentence.
    for c in non_empty:
        if c.startswith("="):
            return False
        if len(c) > 80:
            return False
        if c.endswith("."):
            return False
        # Rough sentence filter: headers don't contain verbs like "is" / "are"
        # embedded in running prose; short phrases are fine.
        if re.search(r"\b(is|are|was|were|should|please)\b", c.lower()):
            return False
    return True


def _is_instruction(label: str) -> bool:
    """True if this looks like prose/an instruction/a URL rather than a field label."""
    if not label:
        return True
    s = label.strip()
    if any(s.startswith(p) for p in INSTRUCTION_PREFIXES):
        return True
    # URLs.
    if re.match(r"^https?://", s):
        return True
    # Long sentences.
    if len(s) > 150:
        return True
    # Sentence-like — contains a period not at the end.
    # (Allow trailing period; flag mid-label periods which signal prose.)
    if re.search(r"\.\s[A-Z]", s):
        return True
    # Continuation of a previous line — starts with lowercase (not a field label).
    if s and s[0].islower():
        return True
    return False


def _col_letter(idx: int) -> str:
    """0-indexed column to Excel A1-style letter (A, B, ..., Z, AA, ...)."""
    idx += 1
    out = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


def _header_index(cells: list[str | None]) -> dict[str, int]:
    return {name: i for i, name in enumerate(cells) if name}


def _is_row_metadata_sheet(ws) -> bool:
    header = [_clean(c.value) for c in ws[1]]
    return ROW_METADATA_REQUIRED_HEADERS.issubset(set(header))


def _normalize_question_number(raw: str) -> str:
    value = _clean(raw)
    if value is None:
        raise ValueError("question number is empty")

    value = re.sub(r"\s+", "", value.upper())
    m = re.match(r"^([A-Z])\.?(.+)$", value)
    if not m:
        raise ValueError(f"unparseable question number: {raw!r}")

    section, rest = m.groups()
    if rest.isdigit() and len(rest) < 3:
        rest = rest.zfill(3)
    return f"{section}.{rest}"


def _subsection_id_from_question_number(raw: str, section_letter: str) -> str:
    value = re.sub(r"\s+", "", str(raw).upper())
    if value.startswith(f"{section_letter}."):
        value = section_letter + value.split(".", 1)[1]
    body = value[1:]

    m = re.match(r"^(\d+)([A-Z]+)?", body)
    if not m:
        return section_letter

    digits, letters = m.groups()
    if letters:
        return f"{section_letter}{digits}{letters}"
    if len(digits) > 2:
        return f"{section_letter}{digits[:-2]}"
    return f"{section_letter}{digits[0]}"


def parse_row_metadata_sheet(ws, section_letter: str) -> dict[str, Any]:
    """Extract structure from the 2024-25 row-metadata sheet format."""
    header = [_clean(c.value) for c in ws[1]]
    index = _header_index(header)
    sheet_name = ws.title

    answer_idx = index["Answer"]
    qnum_idx = index["Question Number"]
    question_idx = index["Question"]
    section_idx = index["Section"]
    subsection_idx = index["Sub-Section"]

    section_title = section_letter
    subsections_by_id: dict[str, dict[str, Any]] = {}
    subsection_order: list[str] = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_qnum = _clean(row[qnum_idx] if qnum_idx < len(row) else None)
        question = _clean(row[question_idx] if question_idx < len(row) else None)
        if not raw_qnum or not question:
            continue

        normalized = _normalize_question_number(raw_qnum)
        if not normalized.startswith(f"{section_letter}."):
            continue

        row_section_title = _clean(row[section_idx] if section_idx < len(row) else None)
        row_subsection_title = _clean(row[subsection_idx] if subsection_idx < len(row) else None)
        if row_section_title and section_title == section_letter:
            section_title = row_section_title

        subsection_id = _subsection_id_from_question_number(raw_qnum, section_letter)
        if subsection_id not in subsections_by_id:
            subsections_by_id[subsection_id] = {
                "id": subsection_id,
                "title": row_subsection_title or subsection_id,
                "questions": [],
            }
            subsection_order.append(subsection_id)

        q: dict[str, Any] = {
            "canonical_question_number": normalized,
            "question_number_raw": raw_qnum,
            "row_label": question,
            "row": row_number,
            "columns": [
                {
                    "header": None,
                    "cell_ref": f"{sheet_name}!{_col_letter(answer_idx)}{row_number}",
                }
            ],
        }
        for source_name, output_name in ROW_METADATA_DIMENSION_HEADERS.items():
            col_idx = index.get(source_name)
            if col_idx is None or col_idx >= len(row):
                continue
            value = _clean(row[col_idx])
            if value is not None:
                q[output_name] = value

        subsections_by_id[subsection_id]["questions"].append(q)

    return {
        "title": section_title,
        "subsections": [subsections_by_id[sub_id] for sub_id in subsection_order],
    }


def parse_section_sheet(ws, section_letter: str) -> list[dict]:
    """Extract subsections + questions from one CDS-X sheet.

    Returns a list of subsection dicts. Each subsection has:
        {
          "id": "B1",
          "title": "Institutional Enrollment - Men and Women",
          "questions": [
              {"row_label": "Degree-seeking, first-time first-year students",
               "row": 13,
               "columns": [
                   {"header": "Males", "cell_ref": "CDS-B!C13"},
                   {"header": "Females", "cell_ref": "CDS-B!D13"},
                   ...
               ]}
          ]
        }
    """
    subsections: list[dict] = []
    current_sub: dict | None = None
    current_header_cells: list[str | None] | None = None
    current_header_row: int | None = None
    section_title: str | None = None
    sheet_name = ws.title

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Pull the first 15 cells, cleaned.
        cells = [_clean(c) for c in (row[:15] if len(row) >= 15 else list(row) + [None] * (15 - len(row)))]

        col_a = cells[0]
        col_b = cells[1]

        # Section title is the first row of the sheet: "B. ENROLLMENT AND PERSISTENCE"
        if section_title is None and col_a and col_a.startswith(f"{section_letter}."):
            section_title = col_a.partition(". ")[2] or col_a
            continue

        # Subsection marker (e.g., "B1") in column A. In templates from
        # 2020-21+, the marker appears once at the top of the block. In
        # 2019-20 the marker appears in col A on every row of the block —
        # so only treat it as a "new subsection" when the id CHANGES.
        is_subsection_row = False
        if col_a:
            m = SUBSECTION_RE.match(col_a)
            if m and m.group(1) == section_letter:
                canonical_id = f"{m.group(1)}{m.group(2)}"
                if current_sub is None or current_sub["id"] != canonical_id:
                    current_sub = {
                        "id": canonical_id,
                        "title": col_b or "",
                        "questions": [],
                    }
                    subsections.append(current_sub)
                    current_header_cells = None
                    current_header_row = None
                    # This row starts a new subsection. If col B is the title,
                    # skip; if col B is a question row data AND this isn't
                    # 2019-20 style, also skip.
                    is_subsection_row = True
                else:
                    # 2019-20 style: col A repeats "B1" on every row of B1.
                    # Normalize col_a to empty so the rest of the logic treats
                    # this as a normal row within B1.
                    col_a = None
                    cells[0] = None
            else:
                # Non-matching col A content (e.g., "A", "B" as alpha sub-labels
                # used inside CDS-I). Don't reset subsection; just skip row.
                continue

        if is_subsection_row:
            continue

        # If we're outside any subsection, skip.
        if current_sub is None:
            continue

        # A header row updates the remembered column headers for subsequent rows.
        if _looks_like_header_row(cells):
            current_header_cells = cells
            current_header_row = i
            continue

        # A question row: column A empty, column B has a non-instruction label.
        if col_a is None and col_b and not _is_instruction(col_b):
            # Numbered prefixes like "1.", "2." in col B are instruction-ish; skip very short fragments.
            if len(col_b) < 3:
                continue

            # Determine which columns carry answer cells. Two cases:
            # 1. Header row remembered — emit one question per header cell in cols 2+.
            # 2. No header — emit a single question with column=null, cell_ref=col C.
            q = {
                "row_label": col_b,
                "row": i,
                "columns": [],
            }
            if current_header_cells:
                for col_idx in range(2, len(current_header_cells)):
                    header = current_header_cells[col_idx]
                    if not header:
                        continue
                    q["columns"].append({
                        "header": header,
                        "cell_ref": f"{sheet_name}!{_col_letter(col_idx)}{i}",
                    })
            else:
                # Default: one answer slot in column C (index 2).
                q["columns"].append({
                    "header": None,
                    "cell_ref": f"{sheet_name}!C{i}",
                })

            if q["columns"]:
                current_sub["questions"].append(q)

    return {"title": section_title or section_letter, "subsections": subsections}


def build_structural_schema(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)

    sections = []
    for name in wb.sheetnames:
        if not name.startswith("CDS-"):
            continue
        if name in ("CDS Definitions",):
            continue
        letter = name.split("-", 1)[1]
        if len(letter) != 1 or not letter.isalpha():
            continue
        ws = wb[name]
        parsed = (
            parse_row_metadata_sheet(ws, letter)
            if _is_row_metadata_sheet(ws)
            else parse_section_sheet(ws, letter)
        )
        sections.append({
            "section": letter,
            "title": parsed["title"],
            "subsections": parsed["subsections"],
        })

    # Count totals.
    subsection_count = sum(len(s["subsections"]) for s in sections)
    question_count = sum(
        len(sub["questions"])
        for s in sections
        for sub in s["subsections"]
    )
    cell_count = sum(
        len(q["columns"])
        for s in sections
        for sub in s["subsections"]
        for q in sub["questions"]
    )

    return {
        "schema_version": _infer_year_from_filename(xlsx_path.name),
        "source_filename": xlsx_path.name,
        "structural": True,
        "source_note": (
            "Extracted from the per-section tabs (CDS-A through CDS-J) of the "
            "commondataset.org Common Data Set Excel template. Row-metadata "
            "templates retain canonical question_numbers when provided by the "
            "source workbook. Visual-layout historical templates require a "
            "separate canonical overlay to cross-reference fields against the "
            "current Answer Sheet schema."
        ),
        "extracted_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "section_count": len(sections),
        "subsection_count": subsection_count,
        "question_row_count": question_count,
        "answer_cell_count": cell_count,
        "sections": sections,
    }


def _infer_year_from_filename(name: str) -> str:
    m = re.search(r"(\d{4})[-_](\d{4})", name)
    if m:
        y1, y2 = m.group(1), m.group(2)
        return f"{y1}-{y2[-2:]}"
    m = re.search(r"(\d{4})[-_](\d{2})(?:\D|$)", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return "unknown"


def main():
    parser = argparse.ArgumentParser(description=__doc__.strip().split("\n\n")[0])
    parser.add_argument("xlsx", type=Path, help="Path to the CDS XLSX template")
    parser.add_argument("output", type=Path, help="Path to write the structural schema JSON")
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} does not exist", file=sys.stderr)
        sys.exit(1)

    schema = build_structural_schema(args.xlsx)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w") as f:
        json.dump(schema, f, indent=2)

    print(
        f"wrote {args.output}\n"
        f"  schema_version:     {schema['schema_version']}\n"
        f"  sections:           {schema['section_count']}\n"
        f"  subsections:        {schema['subsection_count']}\n"
        f"  question rows:      {schema['question_row_count']}\n"
        f"  answer cells total: {schema['answer_cell_count']}"
    )


if __name__ == "__main__":
    main()
