"""
Build a canonical CDS schema JSON from the commondataset.org Excel template.

Reads the `Answer Sheet` tab of the official XLSX template and emits a
structured schema describing every canonical field in the CDS for that year.

The Answer Sheet is maintained by the Common Data Set Initiative and is
the authoritative source of truth for which fields exist, what their IDs
are, and how they map to PDF form-field names. Every version of the CDS
has its own Answer Sheet; this script is expected to run once per template
year.

Usage:
    python tools/schema_builder/build_from_xlsx.py \\
        schemas/templates/cds_2025-26_template.xlsx \\
        schemas/cds_schema_2025_26.json

The output is a JSON document with:
    - metadata (year, source filename, extracted_at)
    - sections: ordered list of section labels seen
    - fields: ordered list of field records, each containing question_number,
      pdf_tag, word_tag, question text, section metadata, value type,
      and a `computed` flag for canonical fields that are derived rather
      than stored in a form field.
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl


ANSWER_SHEET_NAME = "Answer Sheet"
OUTPUT_COLUMNS = {
    "sort_order": "Sort Order",
    "question_number": "Question Number",
    "pdf_tag": "US News PDF Tag",
    "word_tag": "Word Tag",
    "question": "Question",
    "section": "Section",
    "subsection": "Sub-Section",
    "category": "Category",
    "student_group": "Student Group",
    "cohort": "Cohort",
    "residency": "Residency",
    "unit_load": "Unit load",
    "gender": "Gender",
    "value_type": "Value type",
}
REQUIRED_HEADERS = {
    "Question Number",
    "Question",
    "Answer",
    "Section",
    "Sub-Section",
    "Category",
    "Student Group",
    "Cohort",
    "Residency",
    "Unit load",
    "Gender",
    "Value type",
}
OPTIONAL_HEADERS = {"Sort Order", "US News PDF Tag", "Word Tag"}
ALLOWED_HEADERS = REQUIRED_HEADERS | OPTIONAL_HEADERS


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def normalize_question_number(raw: str) -> str:
    value = _clean(raw)
    if value is None:
        raise ValueError("question number is empty")

    m = re.match(r"^([A-Z])\.?(.+)$", value)
    if not m:
        raise ValueError(f"unparseable question number: {raw!r}")

    section, rest = m.groups()
    if rest.isdigit():
        rest = rest.zfill(3)
    return f"{section}.{rest}"


def _header_index(header: list[Optional[str]]) -> dict[str, int]:
    index = {}
    for i, name in enumerate(header):
        if name is None:
            continue
        if name not in ALLOWED_HEADERS:
            raise ValueError(
                f"Unexpected Answer Sheet header at column {i + 1}: {name!r}. "
                f"The template structure may have changed for a new CDS year."
            )
        if name in index:
            raise ValueError(f"Duplicate Answer Sheet header: {name!r}")
        index[name] = i

    missing = sorted(REQUIRED_HEADERS - set(index))
    if missing:
        raise ValueError(
            "Answer Sheet is missing required header(s): " + ", ".join(missing)
        )

    return index


def build_schema(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if ANSWER_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Workbook {xlsx_path} has no sheet named {ANSWER_SHEET_NAME!r}. "
            f"Sheets found: {wb.sheetnames}"
        )
    ws = wb[ANSWER_SHEET_NAME]

    header = [_clean(c.value) for c in ws[1]]
    header_index = _header_index(header)
    has_sort_order = OUTPUT_COLUMNS["sort_order"] in header_index
    has_pdf_tag = OUTPUT_COLUMNS["pdf_tag"] in header_index

    fields = []
    sections_seen = []
    sections_set = set()

    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_ordinal, row in enumerate(rows, start=1):
        if not row:
            continue
        question_number = _clean(_row_value(row, header_index, "question_number"))
        if not question_number:
            continue

        normalized_question_number = normalize_question_number(question_number)
        pdf_tag = _clean(_row_value(row, header_index, "pdf_tag"))
        section = _clean(_row_value(row, header_index, "section"))

        if section and section not in sections_set:
            sections_set.add(section)
            sections_seen.append(section)

        sort_order = _row_value(row, header_index, "sort_order")
        field = {
            "sort_order": int(sort_order) if sort_order is not None else row_ordinal,
            "question_number": normalized_question_number,
            "pdf_tag": pdf_tag,
            "word_tag": _clean(_row_value(row, header_index, "word_tag")),
            "question": _clean(_row_value(row, header_index, "question")),
            "section": section,
            "subsection": _clean(_row_value(row, header_index, "subsection")),
            "category": _clean(_row_value(row, header_index, "category")),
            "student_group": _clean(_row_value(row, header_index, "student_group")),
            "cohort": _clean(_row_value(row, header_index, "cohort")),
            "residency": _clean(_row_value(row, header_index, "residency")),
            "unit_load": _clean(_row_value(row, header_index, "unit_load")),
            "gender": _clean(_row_value(row, header_index, "gender")),
            "value_type": _clean(_row_value(row, header_index, "value_type")),
            "computed": has_pdf_tag and pdf_tag is None,
        }
        fields.append(field)

    if has_sort_order:
        fields.sort(key=lambda f: f["sort_order"] if f["sort_order"] is not None else 10**9)

    return {
        "schema_version": _infer_year_from_filename(xlsx_path.name),
        "source_filename": xlsx_path.name,
        "source_note": (
            "Extracted from the 'Answer Sheet' tab of the commondataset.org "
            "Common Data Set Excel template. This file is derived from the "
            "CDS Initiative's canonical schema and reflects their field "
            "definitions; it is not hand-authored."
        ),
        "extracted_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "field_count": len(fields),
        "sections": sections_seen,
        "fields": fields,
    }


def _row_value(row: tuple, header_index: dict[str, int], column: str):
    header = OUTPUT_COLUMNS[column]
    idx = header_index.get(header)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _infer_year_from_filename(name: str) -> str:
    import re

    m = re.search(r"(\d{4})[-_](\d{4})", name)
    if m:
        y1, y2 = m.group(1), m.group(2)
        return f"{y1}-{y2[-2:]}"
    m = re.search(r"(\d{4})[-_](\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return "unknown"


def main():
    parser = argparse.ArgumentParser(description=__doc__.strip().split("\n\n")[0])
    parser.add_argument("xlsx", type=Path, help="Path to the CDS XLSX template")
    parser.add_argument("output", type=Path, help="Path to write the schema JSON")
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} does not exist", file=sys.stderr)
        sys.exit(1)

    schema = build_schema(args.xlsx)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w") as f:
        json.dump(schema, f, indent=2)

    with_pdf_tag = sum(1 for f in schema["fields"] if f["pdf_tag"])
    computed = sum(1 for f in schema["fields"] if f["computed"])
    print(
        f"wrote {args.output}\n"
        f"  schema_version: {schema['schema_version']}\n"
        f"  total fields:   {schema['field_count']}\n"
        f"  with pdf_tag:   {with_pdf_tag}\n"
        f"  computed:       {computed}\n"
        f"  sections:       {len(schema['sections'])}"
    )


if __name__ == "__main__":
    main()
