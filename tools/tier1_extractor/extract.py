"""
Tier 1 CDS extractor: read filled CDS Excel workbooks.

The CDS Initiative publishes an Excel template with section tabs (CDS-A
through CDS-J) and a hidden lookup layer. Each field has a known cell
position derived from the template's column-AA/AC formulas. Schools that
publish their CDS as a filled copy of this template have values at those
same cell positions — so we can extract deterministically without OCR or
layout parsing.

The extraction strategy:
  1. Parse the CDS template to build a question_number → (sheet, cell) map
     from the column-AA (question number) and column-AC (answer formula)
     on each section tab.
  2. Open the filled XLSX and read the value at each mapped cell position.
  3. Emit canonical JSON keyed by question_number, same shape as Tier 2.

This is the highest-fidelity extraction tier for XLSX sources. Schools
that use the standard template layout get near-100% coverage with zero
ambiguity. Schools with custom layouts will get partial or zero coverage
and should fall through to Tier 4 (Docling) via PDF conversion.

Usage:
    python tools/tier1_extractor/extract.py \\
        path/to/filled.xlsx \\
        schemas/cds_schema_2025_26.json

    # Or with a non-default template:
    python tools/tier1_extractor/extract.py \\
        path/to/filled.xlsx \\
        schemas/cds_schema_2025_26.json \\
        --template schemas/templates/cds_2025-26_template.xlsx

Env: none required (pure local file processing).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


PRODUCER_NAME = "tier1_xlsx"
PRODUCER_VERSION = "0.2.0"
MIN_TEMPLATE_FIELDS_BEFORE_FALLBACK = 25

# Default template path relative to the repo root.
_REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE = _REPO_ROOT / "schemas" / "templates" / "cds_2025-26_template.xlsx"


def normalize_question_number(raw: str) -> str:
    m = re.match(r"^([A-Z])\.?(.+)$", str(raw).strip())
    if not m:
        return str(raw).strip()
    section, rest = m.groups()
    if rest.isdigit():
        rest = rest.zfill(3)
    return f"{section}.{rest}"


def build_cell_map(template_path: Path) -> dict[str, tuple[str, str]]:
    """Parse the CDS template to build {question_number: (sheet_name, cell_ref)}.

    The template has hidden columns on each CDS-* section tab:
      - Column AA (27): question_number (e.g., "A.001", "B.101")
      - Column AC (29): formula referencing the answer cell (e.g., =IF($D$4<>"",$D$4,""))

    We parse the formula to extract the cell reference, giving us a
    deterministic map from question_number to the cell position where
    schools enter their data.
    """
    wb = openpyxl.load_workbook(str(template_path), data_only=False, read_only=True)
    cell_map: dict[str, tuple[str, str]] = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("CDS-"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=27, max_col=29):
            qnum_cell = row[0]  # Column AA
            # Column AB is skipped (index 1)
            formula_cell = row[2]  # Column AC

            qnum = qnum_cell.value
            formula = formula_cell.value
            if not qnum or not formula:
                continue

            # Extract cell reference from IF formula.
            # Patterns: =IF($D$4<>"", $D$4, "")  or  =IF(D4<>"", D4, "")
            # We grab the first column-letter + row-number pair.
            m = re.search(r"[\$]?([A-Z]+)[\$]?(\d+)", str(formula))
            if m:
                cell_ref = f"{m.group(1)}{m.group(2)}"
                cell_map[str(qnum)] = (sheet_name, cell_ref)

    if not cell_map:
        cell_map = build_reduced_layout_cell_map(wb)

    wb.close()
    return cell_map


def build_reduced_layout_cell_map(wb) -> dict[str, tuple[str, str]]:
    """Build a cell map for reduced templates whose section tabs have
    Question Number / Question / Answer columns directly in A/B/C.

    The 2024-25 template has no hidden AA/AC lookup layer. Filled copies
    still put answers in column C next to each canonical question number,
    so column A + C is the deterministic map.
    """
    cell_map: dict[str, tuple[str, str]] = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("CDS-"):
            continue
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=i).value for i in range(1, 4)]
        if header[:3] != ["Question Number", "Question", "Answer"]:
            continue
        for row_idx in range(2, ws.max_row + 1):
            qnum = ws.cell(row=row_idx, column=1).value
            if not qnum:
                continue
            normalized = normalize_question_number(str(qnum))
            cell_map[normalized] = (sheet_name, f"C{row_idx}")
    return cell_map


def load_schema(schema_path: Path) -> dict:
    with schema_path.open() as f:
        return json.load(f)


def extract(xlsx_path: Path, schema: dict, cell_map: dict[str, tuple[str, str]]) -> dict:
    """Extract values from a filled CDS XLSX using the template cell map."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    qnum_to_field = {f["question_number"]: f for f in schema["fields"]}
    values, missing_sheets, empty_count = extract_values_from_cell_map(
        wb,
        cell_map,
        qnum_to_field,
    )
    extraction_layout = "template_cell_map"

    # Some schools publish populated workbooks where the hidden lookup layer
    # itself is filled as a tabular export: Question Number / Question /
    # Answer columns, often in AA/AB/AC. The standard template cell map points
    # to the visible input cells and can produce zero coverage there. If that
    # happens, fall back to the workbook's own answer columns.
    if len(values) < MIN_TEMPLATE_FIELDS_BEFORE_FALLBACK:
        embedded_map = build_embedded_answer_cell_map(wb)
        if embedded_map:
            fallback_values, fallback_missing, fallback_empty = extract_values_from_cell_map(
                wb,
                embedded_map,
                qnum_to_field,
            )
            if len(fallback_values) > len(values):
                values = fallback_values
                missing_sheets = fallback_missing
                empty_count = fallback_empty
                extraction_layout = "embedded_answer_columns"

    wb.close()

    return {
        "producer": PRODUCER_NAME,
        "producer_version": PRODUCER_VERSION,
        "schema_version": schema.get("schema_version"),
        "source_xlsx": xlsx_path.name,
        "extracted_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "stats": {
            "cell_map_fields_total": len(cell_map),
            "schema_fields_populated": len(values),
            "empty_cells": empty_count,
            "missing_sheets": sorted(missing_sheets),
            "extraction_layout": extraction_layout,
        },
        "values": values,
    }


def extract_values_from_cell_map(
    wb,
    cell_map: dict[str, tuple[str, str]],
    qnum_to_field: dict[str, dict],
) -> tuple[dict, set[str], int]:
    available_sheets = set(wb.sheetnames)
    values = {}
    missing_sheets = set()
    empty_count = 0

    for qnum, (sheet_name, cell_ref) in cell_map.items():
        if sheet_name not in available_sheets:
            missing_sheets.add(sheet_name)
            continue

        ws = wb[sheet_name]
        try:
            val = ws[cell_ref].value
        except Exception:
            continue

        if val is None:
            empty_count += 1
            continue

        s = str(val).strip()
        if not s:
            empty_count += 1
            continue

        field = qnum_to_field.get(qnum, {})
        values[qnum] = {
            "value": s,
            "word_tag": field.get("word_tag"),
            "question": field.get("question"),
            "section": field.get("section"),
            "subsection": field.get("subsection"),
            "value_type": field.get("value_type"),
        }

    return values, missing_sheets, empty_count


def build_embedded_answer_cell_map(wb) -> dict[str, tuple[str, str]]:
    """Build qnum → answer-cell map from workbook-native answer columns."""
    cell_map: dict[str, tuple[str, str]] = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("CDS-"):
            continue
        ws = wb[sheet_name]
        header_row, qnum_col, answer_col = find_question_answer_columns(ws)
        if not header_row:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_qnum = ws.cell(row=row_idx, column=qnum_col).value
            if not raw_qnum:
                continue
            qnum = normalize_question_number(str(raw_qnum))
            cell_map[qnum] = (sheet_name, f"{get_column_letter(answer_col)}{row_idx}")
    return cell_map


def find_question_answer_columns(ws) -> tuple[int | None, int | None, int | None]:
    max_col = min(ws.max_column, 40)
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        labels: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            label = re.sub(r"\s+", " ", str(value).strip().lower())
            labels[label] = col_idx
        qnum_col = labels.get("question number") or labels.get("question numer")
        answer_col = labels.get("answer")
        if qnum_col and answer_col:
            return row_idx, qnum_col, answer_col
    return None, None, None


def extract_from_bytes(
    data: bytes,
    schema: dict,
    cell_map: dict[str, tuple[str, str]],
) -> dict:
    """Extract from in-memory bytes (used by the extraction worker)."""
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
        tmp.write(data)
        tmp.flush()
        return extract(Path(tmp.name), schema, cell_map)


def main():
    parser = argparse.ArgumentParser(description=__doc__.strip().split("\n\n")[0])
    parser.add_argument("xlsx", type=Path, help="Path to a filled CDS XLSX")
    parser.add_argument("schema", type=Path, help="Path to a cds_schema JSON")
    parser.add_argument(
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help="Path to the CDS Excel template (default: schemas/templates/cds_2025-26_template.xlsx)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Write output here instead of stdout",
    )
    parser.add_argument(
        "--summary",
        action="store_true",
        help="Print a human-readable summary to stderr",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} does not exist", file=sys.stderr)
        sys.exit(1)
    if not args.schema.exists():
        print(f"error: {args.schema} does not exist", file=sys.stderr)
        sys.exit(1)
    if not args.template.exists():
        print(f"error: template {args.template} does not exist", file=sys.stderr)
        sys.exit(1)

    cell_map = build_cell_map(args.template)
    schema = load_schema(args.schema)
    result = extract(args.xlsx, schema, cell_map)

    payload = json.dumps(result, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload)
    else:
        print(payload)

    if args.summary:
        s = result["stats"]
        print(
            f"\n[{args.xlsx.name}]\n"
            f"  cell map fields:         {s['cell_map_fields_total']}\n"
            f"  schema fields populated: {s['schema_fields_populated']}"
            f" ({s['schema_fields_populated'] * 100 // max(s['cell_map_fields_total'], 1)}%)\n"
            f"  empty cells:             {s['empty_cells']}\n"
            f"  missing sheets:          {s['missing_sheets']}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
