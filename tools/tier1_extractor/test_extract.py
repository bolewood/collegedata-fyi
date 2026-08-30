from __future__ import annotations

import tempfile
import sys
import unittest
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from extract import extract


def _c1_field(
    qnum: str,
    question: str,
    *,
    gender: str = "All",
    residency: str = "All",
    unit_load: str = "All",
) -> dict:
    return {
        "question_number": qnum,
        "word_tag": None,
        "question": question,
        "section": "First-Time, First-Year Admission",
        "subsection": "Applications",
        "value_type": "Number",
        "gender": gender,
        "residency": residency,
        "unit_load": unit_load,
    }


def _schema_2025_c1() -> dict:
    fields = [
        _c1_field("C.101", "Total first-time, first-year males who applied", gender="Males"),
        _c1_field("C.102", "Total first-time, first-year females who applied", gender="Females"),
        _c1_field("C.103", "Total first-time, first-year students of unknown sex who applied", gender="Unknown"),
        _c1_field("C.104", "Total first-time, first-year males who were admitted", gender="Males"),
        _c1_field("C.105", "Total first-time, first-year females who were admitted", gender="Females"),
        _c1_field("C.106", "Total first-time, first-year students of unknown sex who were admitted", gender="Unknown"),
        _c1_field("C.107", "Total first-time, first-year males who enrolled", gender="Males"),
        _c1_field("C.108", "Total first-time, first-year females who enrolled", gender="Females"),
        _c1_field("C.109", "Total first-time, first-year students of unknown sex who enrolled", gender="Unknown"),
        _c1_field(
            "C.110",
            "Total full-time, first-time, first-year males who enrolled",
            gender="Males",
            unit_load="FT",
        ),
        _c1_field("C.116", "Total first-time, first-year students who applied"),
        _c1_field("C.117", "Total first-time, first-year students who were admitted"),
        _c1_field("C.118", "Total first-time, first-year students who enrolled"),
        _c1_field("C.119", "Total first-time, first-year who applied", residency="In-State"),
        _c1_field("C.120", "Total first-time, first-year who were admitted", residency="In-State"),
        _c1_field("C.121", "Total first-time, first-year who enrolled", residency="In-State"),
        _c1_field("C.122", "Total first-time, first-year who applied", residency="Out-of-State"),
        _c1_field("C.123", "Total first-time, first-year who were admitted", residency="Out-of-State"),
        _c1_field("C.124", "Total first-time, first-year who enrolled", residency="Out-of-State"),
        _c1_field("C.125", "Total first-time, first-year who applied", residency="Nonresidents"),
        _c1_field("C.126", "Total first-time, first-year who were admitted", residency="Nonresidents"),
        _c1_field("C.127", "Total first-time, first-year who enrolled", residency="Nonresidents"),
        _c1_field("C.128", "Total first-time, first-year who applied", residency="Unknown"),
        _c1_field("C.129", "Total first-time, first-year who were admitted", residency="Unknown"),
        _c1_field("C.130", "Total first-time, first-year who enrolled", residency="Unknown"),
    ]
    return {"schema_version": "2025-26", "fields": fields}


class Tier1ExtractTests(unittest.TestCase):
    def test_uses_short_section_tab_aliases_for_template_cell_map(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "A"
        ws["D4"] = "Alias University"

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            schema = {
                "schema_version": "2025-26",
                "fields": [{
                    "question_number": "A.101",
                    "word_tag": "institution_name",
                    "question": "Name of College/University:",
                    "section": "General Information",
                    "subsection": "Institutional Contact Information",
                    "value_type": "Text",
                }],
            }
            result = extract(
                Path(tmp.name),
                schema,
                {"A.101": ("CDS-A", "D4")},
            )

        self.assertEqual(result["stats"]["extraction_layout"], "template_cell_map")
        self.assertEqual(result["stats"]["schema_fields_populated"], 1)
        self.assertEqual(result["values"]["A.101"]["value"], "Alias University")
        self.assertEqual(result["stats"]["missing_sheets"], [])

    def test_uses_descriptive_section_tab_aliases_for_c9_recovery(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admissions"
        ws["B10"] = "Percent and number of first-time, first-year students enrolled in Fall 2025 who submitted national standardized (SAT/ACT) test scores."
        ws["B13"] = "Submitting SAT Scores"
        ws["C13"] = 0.55
        ws["D13"] = 2833
        ws["B18"] = "SAT Composite"
        ws["C18"] = 1140
        ws["D18"] = 1250
        ws["E18"] = 1350

        schema = {
            "schema_version": "2025-26",
            "fields": [
                {
                    "question_number": qnum,
                    "word_tag": None,
                    "question": qnum,
                    "section": "First-Time, First-Year Admission",
                    "subsection": "First-time, first-year Profile",
                    "value_type": "Number",
                }
                for qnum in ("C.901", "C.903", "C.905", "C.906", "C.907")
            ],
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            result = extract(
                Path(tmp.name),
                schema,
                {"C.901": ("CDS-C", "C99")},
            )

        self.assertEqual(result["values"]["C.901"]["value"], "0.55")
        self.assertEqual(result["values"]["C.903"]["value"], "2833")
        self.assertEqual(result["values"]["C.905"]["value"], "1140")
        self.assertEqual(result["values"]["C.907"]["value"], "1350")

    def test_falls_back_to_embedded_answer_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CDS-C"
        ws["AA1"] = "Question Number"
        ws["AB1"] = "Question"
        ws["AC1"] = "Answer"
        ws["AA2"] = "C.101"
        ws["AB2"] = "Total first-time, first-year males who applied"
        ws["AC2"] = 1234

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            schema = {
                "schema_version": "2025-26",
                "fields": [{
                    "question_number": "C.101",
                    "word_tag": "c1_male_applicants",
                    "question": "Total first-time, first-year males who applied",
                    "section": "First-Time, First-Year Admission",
                    "subsection": "Applications",
                    "value_type": "Number",
                }],
            }
            result = extract(
                Path(tmp.name),
                schema,
                {"C.101": ("CDS-C", "D4")},
            )

        self.assertEqual(result["stats"]["extraction_layout"], "embedded_answer_columns")
        self.assertEqual(result["stats"]["schema_fields_populated"], 1)
        self.assertEqual(result["values"]["C.101"]["value"], "1234")

    def test_recovers_shifted_c9_academic_profile_rows_by_label(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CDS-C"
        ws["B10"] = "Percent and number of first-time, first-year students enrolled in Fall 2024 who submitted national standardized (SAT/ACT) test scores."
        ws["C12"] = "Percent"
        ws["D12"] = "Number"
        ws["B13"] = "Submitting SAT Scores"
        ws["C13"] = 0.043
        ws["D13"] = 299
        ws["B14"] = "Submitting ACT Scores"
        ws["C14"] = 0.002
        ws["D14"] = 13
        ws["B17"] = "Assessment"
        ws["C17"] = "25th Percentile"
        ws["D17"] = "50th Percentile"
        ws["E17"] = "75th Percentile"
        ws["B18"] = "SAT Composite"
        ws["C18"] = 860
        ws["D18"] = 950
        ws["E18"] = 1050
        ws["B19"] = "SAT Evidence-Based Reading and Writing"
        ws["C19"] = 450
        ws["D19"] = 490
        ws["E19"] = 540
        ws["B20"] = "SAT Math"
        ws["C20"] = 410
        ws["D20"] = 470
        ws["E20"] = 530
        ws["B21"] = "ACT Composite"
        ws["C21"] = 21.5
        ws["D21"] = 25
        ws["E21"] = 29
        ws["A30"] = "C14"
        ws["B30"] = "Application closing date (fall)"
        ws["C30"] = "2025-11-30 00:00:00"

        schema = {
            "schema_version": "2024-25",
            "fields": [
                {
                    "question_number": f"C.{i}",
                    "word_tag": None,
                    "question": f"C.{i}",
                    "section": "First-Time, First-Year Admission",
                    "subsection": "First-time, first-year Profile",
                    "value_type": "Number",
                }
                for i in range(901, 917)
            ],
        }
        cell_map = {
            "C.911": ("CDS-C", "C30"),
            "C.912": ("CDS-C", "C30"),
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            result = extract(Path(tmp.name), schema, cell_map)

        self.assertEqual(result["values"]["C.901"]["value"], "0.043")
        self.assertEqual(result["values"]["C.903"]["value"], "299")
        self.assertEqual(result["values"]["C.911"]["value"], "410")
        self.assertEqual(result["values"]["C.912"]["value"], "470")
        self.assertEqual(result["values"]["C.916"]["value"], "29")
        self.assertEqual(result["stats"]["academic_profile_fields_recovered"], 16)

    def test_recovers_shifted_c1_application_rows_by_label(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admission"
        rows = [
            ("B12", "Total first-time, first-year men who applied", "E12", 13836),
            ("B13", "Total first-time, first-year women who applied", "E13", 17880),
            ("B14", "Total first-time, first-year men who were admitted", "E14", 9914),
            ("B15", "Total first-time, first-year women who were admitted", "E15", 13532),
            ("B16", "Total full-time, first-time, first-year men who enrolled", "E16", 2525),
            ("B17", "Total part-time, first-time, first-year men who enrolled", "E17", 222),
            ("B18", "Total full-time, first-time, first-year women who enrolled", "E18", 3253),
            ("B19", "Total part-time, first-time, first-year women who enrolled", "E19", 218),
            ("B30", "Total first-time, first-year students who applied", "E30", 31716),
            ("B31", "Total first-time, first-year students who were admitted", "E31", 23446),
            ("B32", "Total first-time, first-year students who enrolled", "E32", 6218),
        ]
        for label_cell, label, value_cell, value in rows:
            ws[label_cell] = label
            ws[value_cell] = value

        schema = {
            "schema_version": "2024-25",
            "fields": [
                {
                    "question_number": qnum,
                    "word_tag": None,
                    "question": question,
                    "section": "First-Time, First-Year Admission",
                    "subsection": "Applications",
                    "value_type": "Number",
                    "gender": gender,
                    "residency": "All",
                    "unit_load": unit_load,
                }
                for qnum, gender, unit_load, question in [
                    ("C.101", "Men", "All", "Total first-time, first-year men who applied"),
                    ("C.102", "Women", "All", "Total first-time, first-year women who applied"),
                    ("C.105", "Men", "All", "Total first-time, first-year men who were admitted"),
                    ("C.106", "Women", "All", "Total first-time, first-year women who were admitted"),
                    ("C.109", "Men", "All", "Total full-time, first-time, first-year men who enrolled"),
                    ("C.110", "Men", "PT", "Total part-time, first-time, first-year men who enrolled"),
                    ("C.111", "Women", "FT", "Total full-time, first-time, first-year women who enrolled"),
                    ("C.112", "Women", "PT", "Total part-time, first-time, first-year women who enrolled"),
                    ("C.117", "All", "All", "Total first-time, first-year students who applied"),
                    ("C.118", "All", "All", "Total first-time, first-year students who were admitted"),
                    ("C.119", "All", "All", "Total first-time, first-year students who enrolled"),
                ]
            ],
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            result = extract(Path(tmp.name), schema, {"C.101": ("CDS-C", "D999")})

        self.assertEqual(result["values"]["C.101"]["value"], "13836")
        self.assertEqual(result["values"]["C.102"]["value"], "17880")
        self.assertEqual(result["values"]["C.105"]["value"], "9914")
        self.assertEqual(result["values"]["C.106"]["value"], "13532")
        self.assertEqual(result["values"]["C.109"]["value"], "2525")
        self.assertEqual(result["values"]["C.110"]["value"], "222")
        self.assertEqual(result["values"]["C.111"]["value"], "3253")
        self.assertEqual(result["values"]["C.112"]["value"], "218")
        self.assertEqual(result["values"]["C.117"]["value"], "31716")
        self.assertEqual(result["values"]["C.118"]["value"], "23446")
        self.assertEqual(result["values"]["C.119"]["value"], "6218")
        self.assertEqual(result["stats"]["application_fields_recovered"], 11)

    def test_maps_2025_residency_block_total_last_not_in_state_as_all(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CDS-C"
        ws["B11"] = "Total first-time, first-year males who applied"
        ws["E11"] = 2496
        ws["B16"] = "Total first-time, first-year males who were admitted"
        ws["E16"] = 1424
        ws["B21"] = "Total first-time, first-year males who enrolled"
        ws["E21"] = 264
        ws["B26"] = "Total full-time, first-time, first-year males who enrolled"
        ws["E26"] = 264
        ws["E36"] = "In-State"
        ws["F36"] = "Out-of-State"
        ws["G36"] = "International"
        ws["H36"] = "Unknown"
        ws["I36"] = "Total"
        ws["B37"] = "Total first-time, first-year (degree-seeking) who applied"
        ws["E37"] = 911
        ws["F37"] = 741
        ws["G37"] = 844
        ws["I37"] = 2496
        ws["B38"] = "Total first-time, first-year (degree-seeking) who were admitted"
        ws["E38"] = 657
        ws["F38"] = 519
        ws["G38"] = 248
        ws["I38"] = 1424
        ws["B39"] = "Total first-time, first-year (degree-seeking) who enrolled"
        ws["E39"] = 185
        ws["F39"] = 56
        ws["G39"] = 23
        ws["I39"] = 264

        schema = _schema_2025_c1()
        # Template map already has the correct Total cells; recovery must
        # not zip the in-state column onto C.116/C.117/C.118.
        cell_map = {
            "C.101": ("CDS-C", "E11"),
            "C.104": ("CDS-C", "E16"),
            "C.107": ("CDS-C", "E21"),
            "C.116": ("CDS-C", "I37"),
            "C.117": ("CDS-C", "I38"),
            "C.118": ("CDS-C", "I39"),
            "C.119": ("CDS-C", "E37"),
            "C.122": ("CDS-C", "F37"),
            "C.125": ("CDS-C", "G37"),
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            result = extract(Path(tmp.name), schema, cell_map)

        self.assertEqual(result["values"]["C.101"]["value"], "2496")
        self.assertEqual(result["values"]["C.116"]["value"], "2496")
        self.assertEqual(result["values"]["C.117"]["value"], "1424")
        self.assertEqual(result["values"]["C.118"]["value"], "264")
        self.assertEqual(result["values"]["C.119"]["value"], "911")
        self.assertEqual(result["values"]["C.122"]["value"], "741")
        self.assertEqual(result["values"]["C.125"]["value"], "844")
        self.assertEqual(result["values"]["C.120"]["value"], "657")
        self.assertEqual(result["values"]["C.123"]["value"], "519")
        self.assertEqual(result["values"]["C.126"]["value"], "248")

    def test_maps_2025_residency_block_when_unknown_column_is_filled(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CDS-C"
        ws["B11"] = "Total first-time, first-year males who applied"
        ws["E11"] = 1000
        ws["B12"] = "Total first-time, first-year females who applied"
        ws["E12"] = 2000
        ws["E36"] = "In-State"
        ws["F36"] = "Out-of-State"
        ws["G36"] = "International"
        ws["H36"] = "Unknown"
        ws["I36"] = "Total"
        ws["B37"] = "Total first-time, first-year (degree-seeking) who applied"
        ws["E37"] = 1100
        ws["F37"] = 1200
        ws["G37"] = 400
        ws["H37"] = 300
        ws["I37"] = 3000

        schema = _schema_2025_c1()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            result = extract(Path(tmp.name), schema, {"C.101": ("CDS-C", "Z99")})

        self.assertEqual(result["values"]["C.101"]["value"], "1000")
        self.assertEqual(result["values"]["C.102"]["value"], "2000")
        self.assertEqual(result["values"]["C.116"]["value"], "3000")
        self.assertEqual(result["values"]["C.119"]["value"], "1100")
        self.assertEqual(result["values"]["C.122"]["value"], "1200")
        self.assertEqual(result["values"]["C.125"]["value"], "400")
        self.assertEqual(result["values"]["C.128"]["value"], "300")

    def test_recovers_freshman_c9_header_and_clears_blank_visible_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CDS-C"
        ws["B10"] = "Percent and number of first-time, first-year (freshman) students enrolled in Fall 2024 who submitted national standardized (SAT/ACT) test scores."
        ws["B13"] = "SAT Evidence-Based Reading and Writing"
        ws["C13"] = 600
        ws["D13"] = 660
        ws["E13"] = 720
        ws["B14"] = "SAT Math"
        ws["B20"] = "C10: Class Rank"
        ws["C30"] = "2025-05-01 00:00:00"
        ws["D30"] = 60

        schema = {
            "schema_version": "2024-25",
            "fields": [
                {
                    "question_number": f"C.{i}",
                    "word_tag": None,
                    "question": f"C.{i}",
                    "section": "First-Time, First-Year Admission",
                    "subsection": "First-time, first-year Profile",
                    "value_type": "Number",
                }
                for i in range(908, 914)
            ],
        }
        cell_map = {
            "C.910": ("CDS-C", "D30"),
            "C.911": ("CDS-C", "C30"),
            "C.912": ("CDS-C", "C30"),
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            result = extract(Path(tmp.name), schema, cell_map)

        self.assertEqual(result["values"]["C.908"]["value"], "600")
        self.assertEqual(result["values"]["C.909"]["value"], "660")
        self.assertEqual(result["values"]["C.910"]["value"], "720")
        self.assertNotIn("C.911", result["values"])
        self.assertNotIn("C.912", result["values"])
        self.assertNotIn("C.913", result["values"])
        self.assertEqual(result["stats"]["academic_profile_fields_recovered"], 5)


if __name__ == "__main__":
    unittest.main()
